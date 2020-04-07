import docx
import pandas as pd # can use for excel
from mailmerge import MailMerge
import openpyxl

# can use pandas or openpyxl

wb = openpyxl.load_workbook('Bookmark_sheet.xlsx')
# name the sheet that you are getting data from

sheet = wb["Bookmarks"]
# find the last row for iteration / does not matter if not using for loop

max_column = sheet.max_column

# using mailmerge
template = ("Report_template.docx")
document1 = MailMerge(template)
for i in range(2, max_column):
	document1.merge(
	NAME = str(sheet.cell(row = 2, column = i).value),	
	SERVICENAME = str(sheet.cell(row = 3, column = i).value),
	BRANCH = str(sheet.cell(row = 4, column = i).value),
	Client_Name = str(sheet.cell(row = 5, column = i).value),
	Long_Name = str(sheet.cell(row = 6, column = i).value),
	Inspection_Period = str(sheet.cell(row = 7, column = i).value),
	R_Date = str(sheet.cell(row = 8, column = i).value),
	JBON_Project_Number = str(sheet.cell(row = 9, column = i).value),
	Short_Name = str(sheet.cell(row = 10, column = i).value),
	CompanyName = str(sheet.cell(row = 11, column = i).value),
	Prepared_by = str(sheet.cell(row = 12, column = i).value),
	Checked_by = str(sheet.cell(row = 13, column = i).value),
	Approved_by = str(sheet.cell(row = 14, column = i).value)
	)


# create new file based on the first line in the excel doc
document1.write("NewFile " + str(sheet.cell(row = 1, column = 1).value)+ ".docx")



